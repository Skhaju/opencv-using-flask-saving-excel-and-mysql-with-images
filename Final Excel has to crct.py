from flask import Flask, render_template, Response
import cv2
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import numpy as np
import threading
import mysql.connector

app = Flask(__name__)

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Set up MySQL database connection
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="face"
)
mycursor = mydb.cursor()

def detect_faces(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)
    return faces, gray

def save_to_excel(details, image_size=(300, 300), folder_path="detected_faces"):
    wb = Workbook()
    ws = wb.active
    ws.append(['Date', 'Time', 'Image'])
    
    for timestamp, img in details:
        date_str, time_str = timestamp.split(" ")
        img_path = f"{folder_path}/{timestamp}.jpg"
        cv2.imwrite(img_path, img)
        resized_img = cv2.resize(img, image_size)
        cv2.imwrite(img_path, resized_img)
        img_obj = Image(img_path)
        ws.add_image(img_obj, f'C{len(ws["C"]) + 1}')
        ws['A' + str(len(ws["A"]) + 1)] = date_str
        ws['B' + str(len(ws["B"]) + 1)] = time_str
    
    wb.save('face_recognition_details_with_images.xlsx')

class VideoCamera(object):
    def __init__(self):
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            raise Exception("Could not open video device")

    def __del__(self):
        if self.cap.isOpened():
            self.cap.release()

    def get_frame(self):
        ret, frame = self.cap.read()
        if not ret:
            raise Exception("Failed to grab frame")

        faces, _ = detect_faces(frame)

        for (x, y, w, h) in faces:
            cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
            
            # Check if image with the same timestamp already exists in the database
            mycursor.execute("SELECT * FROM your_table WHERE image_path = %s", (f"{timestamp}.jpg",))
            result = mycursor.fetchone()
            
            if not result:
                save_to_excel([(timestamp, frame)])

                # Save to MySQL database
                image_data = cv2.imencode('.png', frame[y:y+h, x:x+w])[1].tobytes()
                sql = "INSERT INTO your_table (Date, Time, image_path, image_blob) VALUES (%s, %s, %s, %s)"
                val = (timestamp.split(" ")[0], timestamp.split(" ")[1], f"{timestamp}.jpg", image_data)
                mycursor.execute(sql, val)
                mydb.commit()

        ret, jpeg = cv2.imencode('.jpg', frame)
        if not ret:
            raise Exception("Failed to encode frame")

        return jpeg.tobytes()
    

def generate_frames(camera):
    while True:
        frame = camera.get_frame()
        if frame is None:
            break
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(VideoCamera()), mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == '__main__':
    app.run(debug=True)
